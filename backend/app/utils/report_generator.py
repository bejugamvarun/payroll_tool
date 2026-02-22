from typing import Optional
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from app.models.payroll_cycles import PayrollCycle
from app.models.payroll_entries import PayrollEntry
from app.models.salary_components import ComponentType


def generate_salary_statement(
    college_id: int,
    year: int,
    month: int,
    db: Session,
    output_path: str
) -> str:
    """
    Generate salary statement Excel report for a specific college and month.

    Args:
        college_id: College ID
        year: Year
        month: Month
        db: Database session
        output_path: Output file path

    Returns:
        Path to generated Excel file
    """
    # Get payroll cycle
    cycle = db.query(PayrollCycle).filter(
        PayrollCycle.college_id == college_id,
        PayrollCycle.year == year,
        PayrollCycle.month == month
    ).first()

    if not cycle:
        raise ValueError(f"Payroll cycle not found for college {college_id}, {year}-{month}")

    # Get all payroll entries with relationships
    entries = db.query(PayrollEntry).filter(
        PayrollEntry.payroll_cycle_id == cycle.id
    ).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Salary Statement {month:02d}-{year}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:P1')
    title_cell = ws['A1']
    title_cell.value = f"SALARY STATEMENT - {cycle.college.name} - {month:02d}/{year}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Get unique salary components for headers
    earning_components = set()
    deduction_components = set()

    for entry in entries:
        for comp in entry.components:
            if comp.component_type == ComponentType.EARNING:
                earning_components.add(comp.salary_component.name)
            else:
                deduction_components.add(comp.salary_component.name)

    earning_components = sorted(earning_components)
    deduction_components = sorted(deduction_components)

    # Headers
    headers = [
        "S.No",
        "Employee Code",
        "Name",
        "Department",
        "Designation",
        "Working Days",
        "Present Days",
        "Paid Leaves",
        "Comp Leaves",
        "Unpaid Leaves"
    ]

    # Add earning components
    for comp in earning_components:
        headers.append(comp)

    headers.append("Gross Earnings")

    # Add deduction components
    for comp in deduction_components:
        headers.append(comp)

    headers.extend(["Loss of Pay", "Total Deductions", "Net Pay"])

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Write data
    for row_idx, entry in enumerate(entries, start=4):
        employee = entry.employee
        department = employee.department
        designation = employee.designation

        # Create component map
        comp_map = {}
        for comp in entry.components:
            comp_map[comp.salary_component.name] = float(comp.amount)

        row_data = [
            row_idx - 3,  # S.No
            employee.employee_code,
            f"{employee.first_name} {employee.last_name}",
            department.name,
            designation.name,
            cycle.total_working_days,
            float(entry.days_present),
            float(entry.paid_leaves_used),
            float(entry.comp_leaves_used),
            float(entry.unpaid_leaves)
        ]

        # Add earning amounts
        for comp_name in earning_components:
            row_data.append(comp_map.get(comp_name, 0.0))

        row_data.append(float(entry.gross_earnings))

        # Add deduction amounts
        for comp_name in deduction_components:
            row_data.append(comp_map.get(comp_name, 0.0))

        row_data.extend([
            float(entry.loss_of_pay),
            float(entry.total_deductions),
            float(entry.net_pay)
        ])

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border

            # Format numbers
            if col_idx > 5 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    # Add totals row
    total_row = len(entries) + 4
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    # Calculate totals for numeric columns
    for col_idx in range(6, len(headers) + 1):
        if col_idx <= 5:  # Skip non-numeric columns
            continue
        cell = ws.cell(row=total_row, column=col_idx)
        cell.value = f"=SUM({ws.cell(row=4, column=col_idx).coordinate}:{ws.cell(row=total_row-1, column=col_idx).coordinate})"
        cell.font = Font(bold=True)
        cell.border = border
        cell.number_format = '#,##0.00'

    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 15

    # Save workbook
    wb.save(output_path)
    return output_path


def generate_consolidated_report(
    year: int,
    month: int,
    db: Session,
    output_path: str
) -> str:
    """
    Generate consolidated salary report for all colleges.

    Args:
        year: Year
        month: Month
        db: Database session
        output_path: Output file path

    Returns:
        Path to generated Excel file
    """
    # Get all payroll cycles for the month
    cycles = db.query(PayrollCycle).filter(
        PayrollCycle.year == year,
        PayrollCycle.month == month
    ).all()

    if not cycles:
        raise ValueError(f"No payroll cycles found for {year}-{month}")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create a sheet for each college
    for cycle in cycles:
        ws = wb.create_sheet(title=cycle.college.code[:31])  # Excel limit 31 chars

        # Get all entries
        entries = db.query(PayrollEntry).filter(
            PayrollEntry.payroll_cycle_id == cycle.id
        ).all()

        # Similar structure to generate_salary_statement
        # Title
        ws.merge_cells('A1:P1')
        title_cell = ws['A1']
        title_cell.value = f"{cycle.college.name} - {month:02d}/{year}"
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = [
            "S.No", "Employee Code", "Name", "Department", "Designation",
            "Present Days", "Leaves", "Gross", "Deductions", "Net Pay"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data
        for row_idx, entry in enumerate(entries, start=4):
            employee = entry.employee
            row_data = [
                row_idx - 3,
                employee.employee_code,
                f"{employee.first_name} {employee.last_name}",
                employee.department.name,
                employee.designation.name,
                float(entry.days_present),
                float(entry.paid_leaves_used + entry.comp_leaves_used),
                float(entry.gross_earnings),
                float(entry.total_deductions),
                float(entry.net_pay)
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = border
                if col_idx > 5 and isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'

        # Totals
        total_row = len(entries) + 4
        ws.cell(row=total_row, column=1).value = "TOTAL"
        ws.cell(row=total_row, column=1).font = Font(bold=True)

        for col_idx in [6, 7, 8, 9, 10]:  # Numeric columns
            cell = ws.cell(row=total_row, column=col_idx)
            cell.value = f"=SUM({ws.cell(row=4, column=col_idx).coordinate}:{ws.cell(row=total_row-1, column=col_idx).coordinate})"
            cell.font = Font(bold=True)
            cell.border = border
            cell.number_format = '#,##0.00'

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = 15

    # Create summary sheet
    summary_ws = wb.create_sheet(title="Summary", index=0)
    summary_ws.merge_cells('A1:E1')
    summary_ws['A1'] = f"PAYROLL SUMMARY - {month:02d}/{year}"
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A1'].alignment = Alignment(horizontal="center")

    summary_headers = ["College", "Employees", "Gross Amount", "Deductions", "Net Amount"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.border = border

    for row_idx, cycle in enumerate(cycles, start=4):
        entries = db.query(PayrollEntry).filter(
            PayrollEntry.payroll_cycle_id == cycle.id
        ).all()

        total_gross = sum(float(e.gross_earnings) for e in entries)
        total_deductions = sum(float(e.total_deductions) for e in entries)
        total_net = sum(float(e.net_pay) for e in entries)

        row_data = [
            cycle.college.name,
            len(entries),
            total_gross,
            total_deductions,
            total_net
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            if col_idx > 2 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'

    # Grand totals
    grand_total_row = len(cycles) + 4
    summary_ws.cell(row=grand_total_row, column=1).value = "GRAND TOTAL"
    summary_ws.cell(row=grand_total_row, column=1).font = Font(bold=True)

    for col_idx in [2, 3, 4, 5]:
        cell = summary_ws.cell(row=grand_total_row, column=col_idx)
        cell.value = f"=SUM({summary_ws.cell(row=4, column=col_idx).coordinate}:{summary_ws.cell(row=grand_total_row-1, column=col_idx).coordinate})"
        cell.font = Font(bold=True)
        cell.border = border
        if col_idx > 2:
            cell.number_format = '#,##0.00'

    # Adjust column widths
    for col in ['A', 'B', 'C', 'D', 'E']:
        summary_ws.column_dimensions[col].width = 20

    # Save workbook
    wb.save(output_path)
    return output_path
