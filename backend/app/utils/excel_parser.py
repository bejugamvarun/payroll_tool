from typing import Dict, List
from datetime import datetime
import openpyxl
from sqlalchemy.orm import Session
from app.models.employees import Employee
from app.models.attendance_records import AttendanceStatus


def parse_attendance_excel(file_path: str, college_id: int, db: Session) -> Dict:
    """
    Parse attendance Excel file and return structured data.

    Expected format:
    - Row 1: Headers with employee codes or names
    - Column A: Date
    - Columns B onwards: Each column represents an employee
    - Cell values: P (Present), A (Absent), H (Half Day), WW (Weekend Work),
                   HD (Holiday), L (Leave), or empty

    Args:
        file_path: Path to the Excel file
        college_id: College ID for validating employees
        db: Database session

    Returns:
        Dictionary with:
        - records: List of attendance record dictionaries
        - errors: List of error messages
        - count: Number of successful records
    """
    records = []
    errors = []

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet = workbook.active

        # Parse header row to get employee codes
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        employee_codes = []
        employee_map = {}

        # Start from column B (index 1)
        for col_idx in range(1, len(header_row)):
            emp_code = str(header_row[col_idx]).strip() if header_row[col_idx] else None
            if emp_code:
                employee_codes.append(emp_code)

                # Validate employee exists and belongs to college
                employee = db.query(Employee).filter(
                    Employee.employee_code == emp_code,
                    Employee.college_id == college_id,
                    Employee.is_active == True
                ).first()

                if employee:
                    employee_map[col_idx] = employee.id
                else:
                    errors.append(f"Employee code '{emp_code}' not found or not active in college")

        # Parse data rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row[0]:  # Skip if date column is empty
                continue

            # Parse date from first column
            attendance_date = None
            if isinstance(row[0], datetime):
                attendance_date = row[0].date()
            elif isinstance(row[0], str):
                try:
                    attendance_date = datetime.strptime(row[0].strip(), "%Y-%m-%d").date()
                except ValueError:
                    try:
                        attendance_date = datetime.strptime(row[0].strip(), "%d-%m-%Y").date()
                    except ValueError:
                        errors.append(f"Row {row_idx}: Invalid date format '{row[0]}'")
                        continue

            if not attendance_date:
                errors.append(f"Row {row_idx}: Could not parse date")
                continue

            # Parse attendance status for each employee
            for col_idx in range(1, len(row)):
                if col_idx not in employee_map:
                    continue

                cell_value = str(row[col_idx]).strip().upper() if row[col_idx] else ""

                if not cell_value:
                    # Empty cell treated as absent
                    status = AttendanceStatus.ABSENT
                elif cell_value == "P":
                    status = AttendanceStatus.PRESENT
                elif cell_value == "A":
                    status = AttendanceStatus.ABSENT
                elif cell_value == "H":
                    status = AttendanceStatus.HALF_DAY
                elif cell_value == "WW":
                    status = AttendanceStatus.WEEKEND_WORK
                elif cell_value == "HD":
                    status = AttendanceStatus.HOLIDAY
                elif cell_value == "L":
                    status = AttendanceStatus.LEAVE
                else:
                    errors.append(
                        f"Row {row_idx}, Col {col_idx + 1}: Invalid status '{cell_value}'"
                    )
                    continue

                records.append({
                    "employee_id": employee_map[col_idx],
                    "date": attendance_date,
                    "status": status
                })

        workbook.close()

    except Exception as e:
        errors.append(f"Error parsing Excel file: {str(e)}")

    return {
        "records": records,
        "errors": errors,
        "count": len(records)
    }
