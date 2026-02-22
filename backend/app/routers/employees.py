from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from datetime import date, datetime
from app.database import get_db
from app.schemas.employee import EmployeeCreate, EmployeeUpdate, EmployeeResponse, EmployeeBulkCreate, EmployeeBulkUploadResponse
from app.schemas.employee_salary_structure import SalaryStructureResponse, SalaryStructureBulkCreate
from app.schemas.employee_leave_balance import LeaveBalanceResponse
from app.schemas.attendance import AttendanceRecordResponse
from app.models.employees import Employee
from app.models.employee_salary_structures import EmployeeSalaryStructure
from app.models.employee_leave_balances import EmployeeLeaveBalance
from app.models.attendance_records import AttendanceRecord

router = APIRouter(prefix="/employees", tags=["employees"])


def _employee_to_dict(emp: Employee) -> dict:
    return {
        "id": emp.id,
        "employee_code": emp.employee_code,
        "first_name": emp.first_name,
        "last_name": emp.last_name,
        "email": emp.email,
        "phone": emp.phone,
        "college_id": emp.college_id,
        "department_id": emp.department_id,
        "designation_id": emp.designation_id,
        "date_of_joining": emp.date_of_joining,
        "date_of_leaving": emp.date_of_leaving,
        "bank_name": emp.bank_name,
        "bank_account_number": emp.bank_account_number,
        "ifsc_code": emp.ifsc_code,
        "pan_number": emp.pan_number,
        "ctc": emp.ctc,
        "monthly_gross": emp.monthly_gross,
        "is_active": emp.is_active,
        "created_at": emp.created_at,
        "updated_at": emp.updated_at,
        "college_name": emp.college.name if emp.college else None,
        "department_name": emp.department.name if emp.department else None,
        "designation_name": emp.designation.name if emp.designation else None,
    }


@router.get("", response_model=List[EmployeeResponse])
def list_employees(
    college_id: Optional[int] = None,
    department_id: Optional[int] = None,
    is_active: Optional[bool] = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """List all employees with filters"""
    query = db.query(Employee)

    if college_id:
        query = query.filter(Employee.college_id == college_id)
    if department_id:
        query = query.filter(Employee.department_id == department_id)
    if is_active is not None:
        query = query.filter(Employee.is_active == is_active)

    employees = query.offset(skip).limit(limit).all()
    return [_employee_to_dict(emp) for emp in employees]


@router.get("/{employee_id}", response_model=EmployeeResponse)
def get_employee(employee_id: int, db: Session = Depends(get_db)):
    """Get a specific employee by ID"""
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Employee with ID {employee_id} not found"
        )
    return _employee_to_dict(employee)


@router.post("", response_model=EmployeeResponse, status_code=status.HTTP_201_CREATED)
def create_employee(employee: EmployeeCreate, db: Session = Depends(get_db)):
    """Create a new employee"""
    db_employee = Employee(**employee.model_dump())
    db.add(db_employee)
    db.commit()
    db.refresh(db_employee)
    return _employee_to_dict(db_employee)


@router.post("/bulk", response_model=List[EmployeeResponse], status_code=status.HTTP_201_CREATED)
def create_employees_bulk(employees: EmployeeBulkCreate, db: Session = Depends(get_db)):
    """Create multiple employees in bulk"""
    db_employees = [Employee(**employee.model_dump()) for employee in employees.employees]
    db.add_all(db_employees)
    db.commit()
    for emp in db_employees:
        db.refresh(emp)
    return [_employee_to_dict(emp) for emp in db_employees]


@router.post("/upload-excel", response_model=EmployeeBulkUploadResponse)
async def upload_employees_excel(
    college_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Upload employee data from Excel file"""
    import openpyxl
    import io

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    errors = []
    successful = 0
    total = 0

    headers = [cell.value for cell in ws[1]]
    header_map = {str(h).lower().strip(): i for i, h in enumerate(headers) if h}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        total += 1

        try:
            def get_val(key, default=None):
                idx = header_map.get(key)
                if idx is not None and idx < len(row):
                    return row[idx] if row[idx] is not None else default
                return default

            emp = Employee(
                employee_code=str(get_val("employee_code", "")),
                first_name=str(get_val("first_name", "")),
                last_name=str(get_val("last_name", "")),
                email=str(get_val("email", "")),
                phone=str(get_val("phone", "")) if get_val("phone") else None,
                college_id=college_id,
                department_id=int(get_val("department_id", 0)),
                designation_id=int(get_val("designation_id", 0)),
                date_of_joining=get_val("date_of_joining", date.today()),
                bank_name=str(get_val("bank_name", "")) if get_val("bank_name") else None,
                bank_account_number=str(get_val("bank_account_number", "")) if get_val("bank_account_number") else None,
                ifsc_code=str(get_val("ifsc_code", "")) if get_val("ifsc_code") else None,
                pan_number=str(get_val("pan_number", "")) if get_val("pan_number") else None,
                ctc=float(get_val("ctc", 0)),
                monthly_gross=float(get_val("monthly_gross", 0)),
                is_active=True
            )
            db.add(emp)
            db.flush()
            successful += 1
        except Exception as e:
            errors.append({"row": row_idx, "error": str(e)})

    if successful > 0:
        db.commit()

    return EmployeeBulkUploadResponse(
        total=total,
        successful=successful,
        failed=len(errors),
        errors=errors
    )


@router.put("/{employee_id}", response_model=EmployeeResponse)
def update_employee(
    employee_id: int,
    employee_update: EmployeeUpdate,
    db: Session = Depends(get_db)
):
    """Update an employee"""
    db_employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Employee with ID {employee_id} not found"
        )

    update_data = employee_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_employee, key, value)

    db.commit()
    db.refresh(db_employee)
    return _employee_to_dict(db_employee)


@router.delete("/{employee_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_employee(employee_id: int, db: Session = Depends(get_db)):
    """Soft delete an employee"""
    db_employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not db_employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Employee with ID {employee_id} not found"
        )

    db_employee.is_active = False
    db.commit()
    return None


# --- Salary Structure Endpoints ---

@router.get("/{employee_id}/salary-structure", response_model=List[SalaryStructureResponse])
def get_salary_structure(employee_id: int, db: Session = Depends(get_db)):
    """Get salary structure for an employee"""
    structures = db.query(EmployeeSalaryStructure).filter(
        EmployeeSalaryStructure.employee_id == employee_id,
        EmployeeSalaryStructure.effective_to.is_(None) | (EmployeeSalaryStructure.effective_to >= date.today())
    ).all()
    return structures


@router.put("/{employee_id}/salary-structure", response_model=List[SalaryStructureResponse])
def update_salary_structure(
    employee_id: int,
    data: SalaryStructureBulkCreate,
    db: Session = Depends(get_db)
):
    """Update salary structure for an employee (replace all active)"""
    # End-date all current structures
    db.query(EmployeeSalaryStructure).filter(
        EmployeeSalaryStructure.employee_id == employee_id,
        EmployeeSalaryStructure.effective_to.is_(None)
    ).update({"effective_to": date.today()})

    # Create new structures
    new_structures = []
    for s in data.structures:
        structure = EmployeeSalaryStructure(
            employee_id=employee_id,
            salary_component_id=s.salary_component_id,
            amount=s.amount,
            effective_from=s.effective_from,
            effective_to=s.effective_to
        )
        db.add(structure)
        new_structures.append(structure)

    db.commit()
    for s in new_structures:
        db.refresh(s)

    return new_structures


# --- Leave Balance Endpoints ---

@router.get("/{employee_id}/leave-balance", response_model=LeaveBalanceResponse)
def get_leave_balance(
    employee_id: int,
    year: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get leave balance for an employee"""
    if year is None:
        year = date.today().year

    balance = db.query(EmployeeLeaveBalance).filter(
        EmployeeLeaveBalance.employee_id == employee_id,
        EmployeeLeaveBalance.year == year
    ).first()

    if not balance:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No leave balance found for employee {employee_id} for year {year}"
        )

    return balance


# --- Attendance Endpoints ---

@router.get("/{employee_id}/attendance", response_model=List[AttendanceRecordResponse])
def get_employee_attendance(
    employee_id: int,
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Get attendance records for an employee"""
    query = db.query(AttendanceRecord).filter(
        AttendanceRecord.employee_id == employee_id
    )

    if year and month:
        from calendar import monthrange
        start_date = date(year, month, 1)
        _, last_day = monthrange(year, month)
        end_date = date(year, month, last_day)
        query = query.filter(
            AttendanceRecord.date >= start_date,
            AttendanceRecord.date <= end_date
        )

    records = query.order_by(AttendanceRecord.date).all()
    return records
